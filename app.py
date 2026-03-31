"""
국비훈련 통합 관리 시스템
- 연간개설계획 마스터 DB 분석
- 모집현황 텍스트 파싱 (AI 없이 패턴 매칭)
- 과정별 이수자평가 / 비용신청 / 취업성과 / 만족도 추적
- Google Sheets 저장 (팀 공유)
"""
import streamlit as st
import openpyxl
import pandas as pd
import re
import os
import calendar
import difflib
from datetime import datetime, date
from io import BytesIO

# ── Google Sheets (선택적 연동) ──────────────────
def get_gsheet():
    """Google Sheets 연결. secrets 없으면 None 반환"""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_info(
            st.secrets["gcp_service_account"],
            scopes=[
                "https://spreadsheets.google.com/feeds",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        gc = gspread.authorize(creds)
        return gc.open_by_key(st.secrets["sheet_id"])
    except Exception:
        return None

SHEET_COLS = [
    "key","계열","지점","훈련종류","과정명","시작일","종료일","정원",
    "기준주차","확정인원","신청인원","모집률","신청률",
    "개설상태","연기사유","모집비고",
    "이수자평가예정","이수자평가신청일",
    "평가완료","평가완료일","평가비고",
    "비용단위기간",
    "비용신청","비용금액","비용신청일","비용비고",
    "취업_이수자","취업_취업자","취업_조사일","취업비고",
    "만족도점수","만족도조사일","만족도비고",
    "업데이트",
]

# 계열별 지점 목록
SERIES_BRANCHES = {
    "IT":    ["강남","신촌","대구","부산","인천","대전"],
    "컴퓨터":["강남","홍대","부산","부평","대구","대전","광주","수원","구월","일산","울산","노원","분당","종로","안산","천안","안양","청주"],
    "게임":  ["강남","신촌","대구","대전","부평","부산","광주","일산","수원","분당","안산","노원","천안"],
    "뷰티":  ["대구","대전","수원","부산","인천"],
    "요리":  [],
    "승무원":[],
}

@st.cache_data(ttl=30, show_spinner=False)
def load_gsheet_data(_sheet):
    """Google Sheets 전체 데이터 로드"""
    if _sheet is None:
        return {}
    try:
        try:
            ws = _sheet.worksheet("추적DB")
        except Exception:
            ws = _sheet.add_worksheet("추적DB", 1000, len(SHEET_COLS))
            ws.append_row(SHEET_COLS)
            return {}
        records = ws.get_all_records()
        return {r["key"]: r for r in records if r.get("key")}
    except Exception:
        return {}

def save_to_gsheet(sheet, row_dict):
    """Google Sheets에 행 저장 (upsert)"""
    if sheet is None:
        return False
    try:
        try:
            ws = sheet.worksheet("추적DB")
        except Exception:
            ws = sheet.add_worksheet("추적DB", 1000, len(SHEET_COLS))
            ws.append_row(SHEET_COLS)
        all_vals = ws.get_all_values()
        key = row_dict.get("key", "")
        new_row = [str(row_dict.get(c, "")) for c in SHEET_COLS]
        for i, row in enumerate(all_vals[1:], start=2):
            if row and row[0] == key:
                ws.update(f"A{i}", [new_row])
                return True
        ws.append_row(new_row)
        return True
    except Exception as e:
        st.toast(f"Sheets 저장 오류: {e}", icon="⚠️")
        return False

# ── 페이지 설정 ──────────────────────────────────
st.set_page_config(
    page_title="국비훈련 통합 관리",
    page_icon="📋",
    layout="wide",
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap');
html,body,[class*="css"]{font-family:'Noto Sans KR',sans-serif;}
.page-title{font-size:1.55rem;font-weight:700;color:#1a365d;margin-bottom:0.1rem;}
.page-sub{font-size:0.85rem;color:#718096;margin-bottom:0.8rem;}
.kpi-box{background:white;border:1px solid #e2e8f0;border-radius:10px;
         padding:0.85rem 1rem;text-align:center;}
.kpi-num{font-size:1.7rem;font-weight:700;}
.kpi-label{font-size:0.76rem;color:#718096;margin-top:0.1rem;}
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="page-title">📋 국비훈련 통합 관리 시스템</p>', unsafe_allow_html=True)
st.markdown('<p class="page-sub">연간개설계획 · 모집현황 · 이수자평가 · 비용신청 · 취업성과 · 만족도 통합 관리</p>',
            unsafe_allow_html=True)

# ── Google Sheets 연결 상태 ──────────────────────
sheet = get_gsheet()
if sheet:
    st.sidebar.success("✅ Google Sheets 연결됨", icon="🔗")
else:
    st.sidebar.warning("Google Sheets 미연결\n\n데이터가 세션 내에서만 유지됩니다.", icon="⚠️")

# ── 연간개설계획 파일 로드 (자동 또는 업로드) ────────
AUTO_PLAN = "plan.xlsx.xlsx"  # GitHub에 올려둔 파일명

with st.sidebar:
    st.markdown("### 📂 연간개설계획 파일")
    if os.path.exists(AUTO_PLAN):
        st.success(f"✅ plan.xlsx 자동 로드됨", icon="📋")
        plan_file = AUTO_PLAN
        # 다른 파일로 교체 원할 때
        override = st.file_uploader("다른 파일로 교체", type=["xlsx","XLSX"])
        if override:
            plan_file = override
    else:
        plan_file = st.file_uploader(
            "엑셀 업로드 (.xlsx)",
            type=["xlsx", "XLSX"],
        )

if not plan_file:
    st.info("👈 왼쪽 사이드바에서 연간개설계획 엑셀 파일을 업로드해주세요.")
    st.stop()

# ── 헬퍼 함수 ────────────────────────────────────
def get_status(s, e, today):
    if not isinstance(s, datetime) or not isinstance(e, datetime):
        return "정보없음"
    return "완료" if e < today else ("진행중" if s <= today else "예정")

def get_venue(v):
    if not v or str(v).strip() in ("", "None"):
        return "미확인"
    u = str(v).strip().upper()
    return "미확보" if u == "N" else ("확보" if u in ["O","○","◯"] else f"확보({v})")

def classify_reason(r):
    if not r: return "기타"
    r = str(r)
    if "강의장" in r: return "강의장 부족"
    if "모집" in r and ("저조" in r or "부족" in r): return "모집률 저조"
    if "미개설" in r: return "미개설"
    if "강사" in r: return "강사 문제"
    if "효율" in r or "단가" in r: return "비용/효율성"
    if "직종" in r: return "직종 중복/조정"
    return "기타"

def course_key(지점, 과정명, 회차="1"):
    return f"{지점}|{과정명}|{회차}"

def fmt_mmdd(date_str):
    """날짜 문자열에서 MM/DD 추출 (다양한 형식 처리)"""
    s = str(date_str).strip()
    # 표준: "2026-09-17" or "2026-09-17 00:00:00"
    m = re.search(r"\d{4}[-./](\d{2})[-./](\d{2})", s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    # 붙여쓰기: "20260917"
    m2 = re.match(r"\d{4}(\d{2})(\d{2})$", s)
    if m2:
        return f"{m2.group(1)}/{m2.group(2)}"
    # "2026-0917" (두 번째 구분자 누락)
    m3 = re.match(r"\d{4}-(\d{2})(\d{2})$", s)
    if m3:
        return f"{m3.group(1)}/{m3.group(2)}"
    return s[:5] if len(s) >= 5 else s

def calc_단위기간(시작일_str, 종료일_str):
    """개강일 기준 월별 단위기간 리스트 반환"""
    try:
        s = datetime.strptime(str(시작일_str)[:10], "%Y-%m-%d").date()
        e = datetime.strptime(str(종료일_str)[:10], "%Y-%m-%d").date()
    except Exception:
        return []
    periods, cur, 회차 = [], s, 1
    while cur <= e:
        last = calendar.monthrange(cur.year, cur.month)[1]
        end = cur.replace(day=last)
        if end > e:
            end = e
        periods.append({
            "회차": 회차,
            "시작": cur.strftime("%Y.%m.%d"),
            "종료": end.strftime("%Y.%m.%d"),
            "key": f"{회차}",
        })
        cur = (cur.replace(day=28) + __import__('datetime').timedelta(days=4)).replace(day=1)
        회차 += 1
    return periods

def parse_비용단위기간(raw):
    """저장된 단위기간 문자열 → dict 파싱
    포맷: '1|완료|1500000;2|미신청|0'
    """
    result = {}
    if not raw:
        return result
    for seg in str(raw).split(";"):
        parts = seg.split("|")
        if len(parts) == 3:
            result[parts[0]] = {"완료": parts[1] == "완료", "금액": int(parts[2] or 0)}
    return result

def serialize_비용단위기간(data):
    """dict → 저장 문자열"""
    return ";".join(f"{k}|{'완료' if v['완료'] else '미신청'}|{v['금액']}"
                    for k, v in sorted(data.items()))

# ── 메신저 파싱 헬퍼 함수들 ──────────────────────
STAFF_GSHEET_ID = "15aSWtgjDiKUJPtVwPA0H8De6dlRc6LitFcW8H8KkBJk"

@st.cache_data(ttl=1800, show_spinner=False)
def load_staff_from_gsheets(sheet_id):
    """Google Sheets 공개 CSV에서 담당자 현황 자동 로드
    컬럼 구조: NO, 계열, 지점, 국비(담당자명), 입사일, 근무개월수, 취업(담당자명), ...
    """
    try:
        import urllib.request
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read().decode("utf-8")
        from io import StringIO
        df = pd.read_csv(StringIO(raw), header=None, dtype=str).fillna("")

        # 헤더 행 찾기 (계열, 지점 컬럼이 있는 행)
        header_row = None
        for i, row in df.iterrows():
            vals = list(row.values)
            if "계열" in vals and "지점" in vals:
                header_row = i
                break
        if header_row is None:
            return {"__error__": "헤더(계열/지점 컬럼)를 찾을 수 없습니다."}

        headers = list(df.iloc[header_row])
        def fc(name):
            for j, h in enumerate(headers):
                if name == str(h).strip(): return j
            return -1

        계열_col = fc("계열")
        지점_col  = fc("지점")
        국비_col  = fc("국비")   # 국비 담당자명
        취업_col  = fc("취업")   # 취업 담당자명

        def extract_name(cell):
            """'김남희 대리', '장연정 사원\n (육아휴직...)' 등에서 이름만 추출"""
            cell = str(cell).strip().split("\n")[0].strip()
            parts = cell.split()
            if parts:
                return parts[0]  # 첫 단어 = 이름
            return ""

        staff_map = {}
        cur_계열 = ""
        cur_지점 = ""

        for i in range(header_row + 1, len(df)):
            row = df.iloc[i]
            계열_v = row.iloc[계열_col].strip() if 계열_col >= 0 else ""
            지점_v = row.iloc[지점_col].strip() if 지점_col >= 0 else ""

            # 계열/지점 이어받기 (빈 행은 이전 값 유지)
            if 계열_v and 계열_v not in ("nan","합계","소계","계열"):
                cur_계열 = 계열_v
            if 지점_v and 지점_v not in ("nan","합계","소계","지점"):
                cur_지점 = 지점_v

            if not cur_지점 or not cur_계열:
                continue

            # 국비 담당자
            if 국비_col >= 0:
                name = extract_name(row.iloc[국비_col])
                if name and name not in ("nan","국비","-"):
                    staff_map[name] = (cur_계열, cur_지점)

            # 취업 담당자
            if 취업_col >= 0:
                name = extract_name(row.iloc[취업_col])
                if name and name not in ("nan","취업","-"):
                    staff_map[name] = (cur_계열, cur_지점)

        return staff_map
    except Exception as e:
        return {"__error__": str(e)}

@st.cache_data(show_spinner=False)
def parse_staff_file(file_bytes):
    """지점 담당자 현황 파싱 → {이름: (계열, 지점)} 매핑"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    staff_map = {}
    cur_dept = ""
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(row):
            continue
        if row[1]:
            cur_dept = str(row[1]).strip()
        지역 = str(row[2] or "").strip()
        성명 = str(row[3] or "").strip()
        재직 = str(row[5] or "").strip().upper()
        if 성명 and 지역 and 재직 != "X":
            staff_map[성명] = (cur_dept, 지역)
        담당자 = str(row[6] or "").strip()
        담당재직 = str(row[8] or "").strip().upper()
        if 담당자 and 지역 and 담당재직 != "X":
            staff_map[담당자] = (cur_dept, 지역)
    return staff_map

def extract_branch_from_greeting(text):
    """인사말에서 '계열+지점' 패턴 추출. 예: 'IT인천 모집현황' → ('IT','인천')"""
    series_list = sorted(SERIES_BRANCHES.keys(), key=len, reverse=True)
    all_branches = sorted(set(b for bl in SERIES_BRANCHES.values() for b in bl), key=len, reverse=True)
    for ser in series_list:
        for br in all_branches:
            # 붙여쓰기(IT인천) 또는 공백(승무원 강남) 모두 지원
            if (ser + br) in text or (ser + " " + br) in text:
                return (ser, br)
    for ser in series_list:
        if ser in text and ("보고" in text or "현황" in text):
            # 지점명 단독 추출 시도
            for br in all_branches:
                if br in text:
                    return (ser, br)
            return (ser, "")
    return ("", "")

def extract_number(text_val):
    """'13명', '(HRD...) : 9명', '00명' 등에서 숫자 추출"""
    if not text_val:
        return 0
    text_val = str(text_val)
    m = re.search(r'[：:]\s*(\d+)', text_val)
    if m:
        return int(m.group(1))
    nums = re.findall(r'\d+', text_val)
    if nums:
        return int(nums[-1])
    return 0

def parse_date_range(text_val):
    """기간 텍스트 → (시작일, 종료일) YYYY-MM-DD"""
    if not text_val:
        return "", ""
    t = str(text_val).strip()
    # 4자리 연도: 2026-03-20 ~ 2026-08-21 or 2026.03.20.~2026.08.21. (trailing dot 허용)
    m = re.search(r'(\d{4}[-./]\d{1,2}[-./]\d{1,2})\.?\s*~\s*(\d{4}[-./]\d{1,2}[-./]\d{1,2})', t)
    if m:
        def norm(d): return re.sub(r'[./]', '-', d)
        return norm(m.group(1)), norm(m.group(2))
    # 2자리 연도: 26.03.19~26.09.28
    m = re.search(r'(\d{2}[./]\d{2}[./]\d{2})\s*~\s*(\d{2}[./]\d{2}[./]\d{2})', t)
    if m:
        def expand(d):
            p = re.split(r'[./]', d)
            return f"20{p[0]}-{p[1]}-{p[2]}"
        return expand(m.group(1)), expand(m.group(2))
    return "", ""

def split_course_blocks(text):
    """메신저 텍스트에서 과정 블록 분리"""
    # 훈련과정명 / 과정명 모두 지원, 앞에 - · * 숫자. [ 등 다양한 접두사 처리
    course_name_pat = r'(?:훈련\s*)?(?:과\s*정\s*명|과정명)'
    prefix = r'(?:\d+\s*[.·]\s*|[-*·▶►▷]\s*|\[\s*)?'  # 숫자., -, *, ▶, [, 없음
    pat = re.compile(
        r'(?:^|\n)\s*' + prefix + course_name_pat + r'[^:\：]*[:\：]',
        re.MULTILINE
    )
    positions = [m.start() for m in pat.finditer(text)]
    if not positions:
        return []
    blocks = []
    for i, pos in enumerate(positions):
        end = positions[i+1] if i+1 < len(positions) else len(text)
        blocks.append(text[pos:end].strip())
    return blocks

# 필드 레이블 정규화 테이블
# 키: 정규식 패턴(띄어쓰기 무관), 값: 내부 필드명
FIELD_LABEL_MAP = [
    # (정규식패턴,                              내부키,       처리방식)
    (r'(?:훈련\s*)?(?:과\s*정\s*명)',           "과정명",     "text"),
    (r'(?:훈련\s*)?기\s*간',                    "기간",       "date"),
    (r'훈련\s*시\s*간',                         "훈련시간",   "text"),
    (r'훈련\s*일\s*수',                         "훈련일수_m", "text"),  # 메신저 훈련일수(참고용)
    (r'강\s*의\s*[장실]',                       "강의장",     "text"),
    (r'(?:모\s*집\s*인\s*원|정\s*원)',          "모집인원",   "number"),
    (r'신\s*청\s*인\s*원',                      "신청인원",   "number"),
    (r'확\s*정\s*인\s*원',                      "확정인원",   "number"),
    (r'(?:HRD[^:\：]*)?신\s*청',               "신청인원",   "number"),  # HRD신청 등
    (r'회\s*차',                                "회차",       "text"),
]

def _norm_label(label):
    """레이블에서 공백/특수문자 제거해 정규화"""
    return re.sub(r'[\s\(\)\[\]△▶►▷·*\-]', '', label)

def parse_one_course(block):
    """단일 과정 블록 → dict (레이블 테이블 기반, 띄어쓰기/특수문자 무관)"""
    result = {
        "과정명": "", "시작일": "", "종료일": "",
        "훈련시간": "", "훈련일수_m": "", "강의장": "",
        "모집인원": 0, "신청인원": 0, "확정인원": 0, "회차": "",
    }
    for raw_line in block.split('\n'):
        # 줄 앞 접두사 제거 (▶ - * · 숫자. 등)
        line = re.sub(r'^[\s▶►▷\-*·]+', '', raw_line).strip()
        if not line:
            continue
        # 콜론 분리: "레이블 : 값" 또는 "레이블: 값"
        sep = re.search(r'[:\：]', line)
        if not sep:
            continue
        label_raw = line[:sep.start()].strip()
        value_raw = line[sep.end():].strip()
        # 괄호 속 보충설명 제거 (예: "훈련기간 (총일수)")
        label_clean = re.sub(r'\([^)]*\)', '', label_raw).strip()

        matched_key = None
        matched_type = None
        for pat, key, typ in FIELD_LABEL_MAP:
            if re.fullmatch(pat + r'.*', label_clean, re.IGNORECASE):
                matched_key = key
                matched_type = typ
                break

        if matched_key is None:
            continue

        if matched_type == "text":
            val = re.sub(r'^\[', '', value_raw).rstrip(']').strip()
            result[matched_key] = val
        elif matched_type == "date":
            s, e = parse_date_range(value_raw)
            result["시작일"] = s
            result["종료일"] = e
        elif matched_type == "number":
            result[matched_key] = extract_number(value_raw)

    return result

def fuzzy_match_plan(course_name, branch, plan_courses):
    """과정명+지점으로 연간개설계획 매칭"""
    if not course_name:
        return None
    def normalize(s):
        return re.sub(r'[\s\(\)\[\]&·]', '', str(s)).lower()
    norm_q = normalize(course_name)
    candidates = [c for c in plan_courses if branch and branch in (c.get("지점",""))] if branch else plan_courses
    if not candidates:
        candidates = plan_courses
    for c in candidates:
        if normalize(c["과정명"]) == norm_q:
            return c
    best, best_score = None, 0
    for c in candidates:
        score = difflib.SequenceMatcher(None, norm_q, normalize(c["과정명"])).ratio()
        if score > best_score and score > 0.55:
            best, best_score = c, score
    return best

def parse_messenger_all(text, staff_map, plan_courses):
    """전체 메신저 텍스트 파싱 → 과정 데이터 리스트"""
    results = []
    name_pat = re.compile(r'^([가-힣]{2,4})$', re.MULTILINE)
    matches = list(name_pat.finditer(text))
    if not matches:
        blocks = [("", text)]
    else:
        blocks = []
        for i, m in enumerate(matches):
            end = matches[i+1].start() if i+1 < len(matches) else len(text)
            blocks.append((m.group(1), text[m.start():end]))
    for name, block in blocks:
        계열, 지점 = extract_branch_from_greeting(block)
        if name in staff_map:
            s_계, s_지 = staff_map[name]
            if not 계열: 계열 = s_계
            if not 지점: 지점 = s_지
        for cb in split_course_blocks(block):
            c = parse_one_course(cb)
            if not c["과정명"]:
                continue
            plan = fuzzy_match_plan(c["과정명"], 지점, plan_courses)
            # 연간개설계획 데이터 우선, 없을 때만 메신저 값 사용
            plan_정원 = int(plan.get("정원",0) or 0) if plan else 0
            정원 = plan_정원 or c["모집인원"]
            확정 = c["확정인원"]
            신청 = c["신청인원"]
            results.append({
                "보고자": name,
                "계열": 계열 or (plan.get("계열","") if plan else ""),
                "지점": 지점 or (plan.get("지점","") if plan else ""),
                "훈련종류": plan.get("훈련종류","") if plan else "",
                "과정명": (plan.get("과정명","") if plan else "") or c["과정명"],
                "시작일": (plan.get("시작일","") if plan else "") or c["시작일"],
                "종료일": (plan.get("종료일","") if plan else "") or c["종료일"],
                "훈련일수": plan.get("훈련일수","") if plan else "",
                "훈련시간": (plan.get("훈련시간","") if plan else "") or c["훈련시간"],
                "정원": 정원,
                "확정인원": 확정,
                "신청인원": 신청,
                "모집률(%)": round(확정/정원*100, 1) if 정원 > 0 else 0,
                "신청률(%)": round(신청/정원*100, 1) if 정원 > 0 else 0,
                "강의장": c["강의장"],
                "매칭과정명": plan.get("과정명","") if plan else "",
                "비고": "",
            })
    return results

def export_messenger_excel(rows, week_label):
    """파싱 결과를 엑셀 보고 양식으로 출력"""
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "모집현황"
    ws.merge_cells("A1:O1")
    ws["A1"] = f"26년 {week_label} 모집현황"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    headers = ["계열","지점","훈련종류","훈련과정명","훈련일(일)","훈련시간","시작일","종료일","정원","확정인원","신청인원","모집률","신청률","강의장","비고"]
    col_widths = [8, 8, 12, 50, 8, 18, 12, 12, 6, 8, 8, 8, 8, 12, 20]
    for j, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=2, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1a365d")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[2].height = 20
    for i, r in enumerate(rows, 3):
        ws.cell(row=i, column=1, value=r.get("계열",""))
        ws.cell(row=i, column=2, value=r.get("지점",""))
        ws.cell(row=i, column=3, value=r.get("훈련종류",""))
        ws.cell(row=i, column=4, value=r.get("과정명",""))
        ws.cell(row=i, column=5, value=r.get("훈련일수",""))
        ws.cell(row=i, column=6, value=r.get("훈련시간",""))
        ws.cell(row=i, column=7, value=r.get("시작일",""))
        ws.cell(row=i, column=8, value=r.get("종료일",""))
        ws.cell(row=i, column=9, value=r.get("정원",0))
        ws.cell(row=i, column=10, value=r.get("확정인원",0))
        ws.cell(row=i, column=11, value=r.get("신청인원",0))
        mr = r.get("모집률(%)",0)
        sr = r.get("신청률(%)",0)
        c_mr = ws.cell(row=i, column=12, value=round(mr/100,4))
        c_sr = ws.cell(row=i, column=13, value=round(sr/100,4))
        c_mr.number_format = "0.0%"
        c_sr.number_format = "0.0%"
        ws.cell(row=i, column=14, value=r.get("강의장",""))
        ws.cell(row=i, column=15, value=r.get("비고",""))
        # 모집률 낮으면 빨간색
        if r.get("정원",0) > 0 and mr < 65:
            c_mr.font = Font(color="C53030", bold=True)
        if r.get("정원",0) > 0 and sr < 70:
            c_sr.font = Font(color="C53030", bold=True)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ── 마스터 DB 파싱 ────────────────────────────────
@st.cache_data(show_spinner="파일 읽는 중...")
def parse_plan(file_bytes):
    wb    = openpyxl.load_workbook(BytesIO(file_bytes))
    today = datetime.now()
    ws1   = wb["1.연간개설계획"]
    courses = []
    for row in ws1.iter_rows(min_row=6, values_only=True):
        if not row[1]:
            continue
        s, e, v = row[9], row[10], row[16]
        courses.append({
            "연번": row[1], "계열": str(row[2] or ""), "지점": str(row[3] or ""),
            "훈련종류": str(row[4] or ""), "직종": str(row[5] or ""),
            "과정명": str(row[6] or ""), "운영회차": str(row[7] or "1"),
            "시작일": s.strftime("%Y-%m-%d") if isinstance(s, datetime) else str(s or ""),
            "종료일": e.strftime("%Y-%m-%d") if isinstance(e, datetime) else str(e or ""),
            "훈련일수": row[11], "훈련시간": row[12], "정원": int(row[15] or 0),
            "강의장": str(v or ""), "강사": str(row[17] or ""), "비고": str(row[18] or ""),
            "진행상태": get_status(s, e, today),
            "강의장상태": get_venue(v),
        })
    ws2 = wb["2.반납과정"]
    returns = []
    for row in ws2.iter_rows(min_row=6, values_only=True):
        if not row[1]:
            continue
        r = row[11]
        returns.append({
            "연번": row[1], "계열": str(row[2] or ""), "지점": str(row[3] or ""),
            "훈련종류": str(row[4] or ""), "직종": str(row[5] or ""),
            "과정명": str(row[6] or ""),
            "반납사유": str(r or ""), "사유분류": classify_reason(r),
        })
    return courses, returns

if isinstance(plan_file, str):
    with open(plan_file, "rb") as f:
        file_bytes = f.read()
else:
    file_bytes = plan_file.read()
courses, returns = parse_plan(file_bytes)

# ── 모집현황 누적 파일 파싱 함수 ──────────────────
@st.cache_data(show_spinner=False)
def parse_recruit_sheet(file_bytes, sheet_name):
    """모집현황 엑셀의 특정 시트 파싱 (26년/25년/24년 등)
    Row1=집계, Row2=헤더, Row3+=데이터
    컬럼: 월,계열,지점,훈련종류,직종,회차,과정명,훈련일,훈련시간,시작일,종료일,정원,확정인원,신청인원,모집률,신청률
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[6]:   # 과정명 없으면 skip
            continue
        def v(x): return x if x is not None else ""
        def f(x):
            try: return round(float(x), 4)
            except: return 0.0
        def i(x):
            try: return int(float(x)) if x not in (None, "", " ") else 0
            except: return 0
        s = row[9]
        e = row[10]
        rows.append({
            "월":      str(v(row[0])),
            "계열":    str(v(row[1])),
            "지점":    str(v(row[2])),
            "훈련종류":str(v(row[3])),
            "직종":    str(v(row[4])),
            "회차":    str(v(row[5]) or "1"),
            "과정명":  str(v(row[6])),
            "훈련일":  v(row[7]),
            "훈련시간":v(row[8]),
            "시작일":  s.strftime("%Y-%m-%d") if isinstance(s, datetime) else str(s or ""),
            "종료일":  e.strftime("%Y-%m-%d") if isinstance(e, datetime) else str(e or ""),
            "정원":    i(row[11]),
            "확정인원":i(row[12]),
            "신청인원":i(row[13]),
            "모집률":  f(row[14]),
            "신청률":  f(row[15]),
        })
    return rows

@st.cache_data(show_spinner=False)
def parse_recruit_summary(file_bytes):
    """20~25년 모집 누적 시트에서 연도별 월별 집계 파싱"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    sheet_name = "20~25년 모집 누적"
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    # 구조: 각 연도 블록이 반복 (연도헤더, 구분, 1월~12월)
    # B열=구분명, C~N열=1월~12월
    summary = {}
    cur_year = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        b = str(row[1] or "").strip()
        if "년도" in b or ("20" in b and "년" in b):
            m = re.search(r"(20\d{2})", b)
            if m: cur_year = int(m.group(1))
        elif cur_year and b in ("과정수", "정원", "확정인원", "신청인원", "모집률", "신청률"):
            if cur_year not in summary:
                summary[cur_year] = {}
            vals = [row[i] for i in range(2, 14)]  # 1월~12월
            def safe_float(x):
                try: return float(x) if x is not None else 0.0
                except: return 0.0
            summary[cur_year][b] = [safe_float(x) for x in vals]
    return summary

# ── 사이드바: 모집현황 파일 업로드 ──────────────
with st.sidebar:
    st.markdown("### 📊 모집현황 누적 파일")
    AUTO_RECRUIT = "recruit.xlsx"
    if os.path.exists(AUTO_RECRUIT):
        st.success("✅ recruit.xlsx 자동 로드됨", icon="📊")
        recruit_file = AUTO_RECRUIT
        r_override = st.file_uploader("모집현황 파일 교체", type=["xlsx","XLSX"], key="rec_up")
        if r_override: recruit_file = r_override
    else:
        recruit_file = st.file_uploader(
            "모집현황 누적 엑셀 업로드\n(★모집현황 누적본)", type=["xlsx","XLSX"], key="rec_up"
        )

recruit_bytes = None
if recruit_file:
    if isinstance(recruit_file, str):
        with open(recruit_file,"rb") as f: recruit_bytes = f.read()
    else:
        recruit_bytes = recruit_file.read()

# ── 담당자 현황: Google Sheets 자동 로드 ─────────
staff_map = load_staff_from_gsheets(STAFF_GSHEET_ID)

with st.sidebar:
    st.markdown("### 👤 지점 담당자 현황")
    _err = staff_map.get("__error__") if isinstance(staff_map, dict) else None
    _real_map = {k: v for k, v in staff_map.items() if k != "__error__"} if isinstance(staff_map, dict) else {}
    if _real_map:
        st.success(f"✅ 담당자 {len(_real_map)}명 자동 로드 (Google Sheets)", icon="👤")
        staff_map = _real_map
    else:
        if _err:
            st.error(f"로드 실패: {_err}", icon="⚠️")
        else:
            st.warning("담당자 데이터 없음 — 파일 직접 업로드해주세요", icon="⚠️")
        staff_map = {}
    # 수동 업로드로 덮어쓰기 가능
    s_override = st.file_uploader("담당자 파일 직접 업로드 (선택)", type=["xlsx","XLSX"], key="staff_up")
    if s_override:
        try:
            staff_map = parse_staff_file(s_override.read())
            st.success(f"✅ 파일에서 {len(staff_map)}명 로드됨", icon="📁")
        except Exception:
            st.error("파일 파싱 실패")

# ── Google Sheets 데이터 로드 (또는 session_state) ──
if sheet:
    db = load_gsheet_data(sheet)
else:
    if "local_db" not in st.session_state:
        st.session_state.local_db = {}
    db = st.session_state.local_db

def save_record(key, record):
    """저장 (Sheets 또는 session_state)"""
    if sheet:
        save_to_gsheet(sheet, record)
        st.cache_data.clear()
    else:
        st.session_state.local_db[key] = record

# ── 상단 KPI ─────────────────────────────────────
total      = len(courses)
in_prog    = sum(1 for c in courses if c["진행상태"] == "진행중")
scheduled  = sum(1 for c in courses if c["진행상태"] == "예정")
confirmed  = sum(1 for r in db.values() if str(r.get("개설상태","")) == "개강확정")
warn_cnt   = sum(1 for r in db.values()
                 if str(r.get("개설상태","")) == "개강확정"
                 and (float(r.get("모집률",1) or 1) < 0.65
                      or float(r.get("신청률",1) or 1) < 0.70))
delay_cnt  = sum(1 for r in db.values() if str(r.get("개설상태","")) == "개강연기")

kpi_cols = st.columns(6)
for col, num, lbl, clr in [
    (kpi_cols[0], total,     "전체 계획과정", "#2b6cb0"),
    (kpi_cols[1], in_prog,   "진행중",       "#276749"),
    (kpi_cols[2], scheduled, "예정",         "#2c5282"),
    (kpi_cols[3], confirmed, "개강확정",     "#276749"),
    (kpi_cols[4], warn_cnt,  "모집경고",     "#e53e3e"),
    (kpi_cols[5], delay_cnt, "개강연기",     "#c05621"),
]:
    with col:
        st.markdown(
            f'<div class="kpi-box"><div class="kpi-num" style="color:{clr}">{num}</div>'
            f'<div class="kpi-label">{lbl}</div></div>',
            unsafe_allow_html=True,
        )

st.markdown("<br>", unsafe_allow_html=True)

# ════════════════════════════════════════════════
# 탭 구성
# ════════════════════════════════════════════════
tab0, tab_msg, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📋 개설 계획",
    "📨 모집현황(메신저 분류)",
    "📊 모집현황 조회",
    "🔍 과정 추적 관리",
    "🔴 반납 분석",
    "🏅 인증평가 현황",
    "📈 연도별 비교",
])

# ══════════════════════════════════════════════
# TAB 0 : 개설 계획
# ══════════════════════════════════════════════
with tab0:
    from collections import defaultdict

    def empty_stat():
        return {"수": 0, "정원": 0, "진행중": 0, "예정": 0, "완료": 0, "미확보": 0}

    def add_stat(d, c):
        d["수"] += 1
        d["정원"] += c["정원"]
        d[c["진행상태"]] = d.get(c["진행상태"], 0) + 1
        if c["강의장상태"] == "미확보":
            d["미확보"] += 1

    s_tot = defaultdict(empty_stat)
    s_br  = defaultdict(lambda: defaultdict(empty_stat))
    for c in courses:
        ser = c["계열"] or "미분류"
        br  = c["지점"] or "미분류"
        add_stat(s_tot[ser], c)
        add_stat(s_br[ser][br], c)

    s_tot = dict(sorted(s_tot.items()))
    s_br  = {k: dict(sorted(v.items())) for k, v in sorted(s_br.items())}

    G = "2fr 0.7fr 0.9fr 0.6fr 0.6fr 0.6fr 0.75fr"

    def th():
        return (
            f"<div style='display:grid;grid-template-columns:{G};background:#1a365d;color:white;"
            f"border-radius:6px 6px 0 0;padding:0.45rem 0.8rem;font-size:0.79rem;font-weight:700;gap:4px'>"
            "<span>구분</span><span style='text-align:center'>과정수</span>"
            "<span style='text-align:center'>정원합계</span><span style='text-align:center'>진행중</span>"
            "<span style='text-align:center'>예정</span><span style='text-align:center'>완료</span>"
            "<span style='text-align:center'>강의장미확보</span></div>"
        )

    def tr(label, d, bg, bold=False, indent=False):
        lc  = "#1a365d" if bold else "#4a5568"
        fw  = "700"     if bold else "400"
        pl  = "1.6rem"  if indent else "0.8rem"
        pre = "└ "      if indent else ""
        wc  = "color:#c05621;font-weight:700;" if d["미확보"] > 0 else ""
        return (
            f"<div style='display:grid;grid-template-columns:{G};background:{bg};"
            f"border:1px solid #e2e8f0;border-top:none;"
            f"padding:0.38rem 0.8rem 0.38rem {pl};font-size:0.81rem;gap:4px;align-items:center'>"
            f"<span style='color:{lc};font-weight:{fw}'>{pre}{label}</span>"
            f"<span style='text-align:center;font-weight:700;color:#2b6cb0'>{d['수']}</span>"
            f"<span style='text-align:center'>{d['정원']:,}명</span>"
            f"<span style='text-align:center;color:#276749;font-weight:600'>{d.get('진행중',0)}</span>"
            f"<span style='text-align:center;color:#2c5282'>{d.get('예정',0)}</span>"
            f"<span style='text-align:center;color:#718096'>{d.get('완료',0)}</span>"
            f"<span style='text-align:center;{wc}'>{d['미확보'] if d['미확보']>0 else '-'}</span>"
            "</div>"
        )

    # DB에서 각 과정의 개설상태 매핑 (지점+과정명으로 조회)
    def get_db_status(c):
        k = course_key(c["지점"], c["과정명"], c.get("운영회차","1"))
        rec = db.get(k)
        if rec:
            return rec.get("개설상태",""), rec.get("확정인원",""), rec.get("신청인원",""), rec.get("모집률","")
        return "", "", "", ""

    STATUS_ICON = {
        "개강확정": "✅", "개강연기": "🔄", "폐강": "✖", "준비중": "⬜"
    }

    # 집계에 개설상태 반영
    def empty_stat2():
        return {"수":0,"정원":0,"진행중":0,"예정":0,"완료":0,"미확보":0,"개강확정":0,"미등록":0}

    s_tot2 = defaultdict(empty_stat2)
    s_br2  = defaultdict(lambda: defaultdict(empty_stat2))
    for c in courses:
        ser = c["계열"] or "미분류"
        br  = c["지점"] or "미분류"
        st2, _, _, _ = get_db_status(c)
        for d in [s_tot2[ser], s_br2[ser][br]]:
            d["수"] += 1
            d["정원"] += c["정원"]
            d[c["진행상태"]] = d.get(c["진행상태"],0) + 1
            if c["강의장상태"] == "미확보": d["미확보"] += 1
            if st2 == "개강확정": d["개강확정"] += 1
            elif not st2: d["미등록"] += 1

    s_tot2 = dict(sorted(s_tot2.items()))
    s_br2  = {k: dict(sorted(v.items())) for k, v in sorted(s_br2.items())}

    G2 = "2fr 0.6fr 0.8fr 0.6fr 0.6fr 0.6fr 0.7fr 0.7fr"

    def th2():
        return (
            f"<div style='display:grid;grid-template-columns:{G2};background:#1a365d;color:white;"
            f"border-radius:6px 6px 0 0;padding:0.45rem 0.8rem;font-size:0.79rem;font-weight:700;gap:4px'>"
            "<span>구분</span><span style='text-align:center'>과정수</span>"
            "<span style='text-align:center'>정원합계</span><span style='text-align:center'>진행중</span>"
            "<span style='text-align:center'>예정</span><span style='text-align:center'>완료</span>"
            "<span style='text-align:center'>✅개강확정</span>"
            "<span style='text-align:center'>❓미등록</span></div>"
        )

    def tr2(label, d, bg, bold=False, indent=False):
        lc = "#1a365d" if bold else "#4a5568"
        fw = "700" if bold else "400"
        pl = "1.6rem" if indent else "0.8rem"
        pre = "└ " if indent else ""
        uc = "color:#e53e3e;font-weight:700;" if d.get("미등록",0) > 0 else "color:#718096;"
        return (
            f"<div style='display:grid;grid-template-columns:{G2};background:{bg};"
            f"border:1px solid #e2e8f0;border-top:none;"
            f"padding:0.38rem 0.8rem 0.38rem {pl};font-size:0.81rem;gap:4px;align-items:center'>"
            f"<span style='color:{lc};font-weight:{fw}'>{pre}{label}</span>"
            f"<span style='text-align:center;font-weight:700;color:#2b6cb0'>{d['수']}</span>"
            f"<span style='text-align:center'>{d['정원']:,}명</span>"
            f"<span style='text-align:center;color:#276749;font-weight:600'>{d.get('진행중',0)}</span>"
            f"<span style='text-align:center;color:#2c5282'>{d.get('예정',0)}</span>"
            f"<span style='text-align:center;color:#718096'>{d.get('완료',0)}</span>"
            f"<span style='text-align:center;color:#276749;font-weight:700'>{d.get('개강확정',0)}</span>"
            f"<span style='text-align:center;{uc}'>{d.get('미등록',0) if d.get('미등록',0)>0 else '-'}</span>"
            "</div>"
        )

    st.markdown("**📂 계열 > 지점별 현황**")
    st.markdown(th2(), unsafe_allow_html=True)
    ri = 0
    for ser, sv in s_tot2.items():
        st.markdown(tr2(ser, sv, "#ebf4ff", bold=True), unsafe_allow_html=True)
        for br, bv in s_br2[ser].items():
            st.markdown(tr2(br, bv, "#fff" if ri%2==0 else "#f7fafc", indent=True),
                        unsafe_allow_html=True)
            ri += 1
    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("**📄 과정 목록**")
    cf1, cf2, cf3, cf4 = st.columns(4)
    with cf1:
        sf = st.multiselect("진행상태", ["진행중","예정","완료","정보없음"],
                            default=["진행중","예정"])
    with cf2:
        vf = st.multiselect("강의장", ["확보","미확보","미확인"],
                            default=["확보","미확보","미확인"])
    with cf3:
        bf = st.multiselect("지점", sorted(set(c["지점"] for c in courses if c["지점"])))
    with cf4:
        stf0 = st.multiselect("개설상태", ["✅ 개강확정","🔄 개강연기","✖ 폐강","⬜ 준비중","❓ 미등록"],
                              default=[])

    fc = [c for c in courses
          if c["진행상태"] in sf and c["강의장상태"] in vf
          and (not bf or c["지점"] in bf)]

    # 개설상태 필터 적용
    if stf0:
        def match_stf(c):
            st2, _, _, _ = get_db_status(c)
            icon = STATUS_ICON.get(st2, "❓")
            label = f"{icon} {st2}" if st2 else "❓ 미등록"
            return label in stf0
        fc = [c for c in fc if match_stf(c)]

    st.caption(f"{len(fc)}개 과정")

    if fc:
        rows0 = []
        for c in fc:
            st2, 확정, 신청, mr = get_db_status(c)
            icon = STATUS_ICON.get(st2, "❓")
            개설라벨 = f"{icon} {st2}" if st2 else "❓ 미등록"
            정원 = c["정원"] or 1
            mr_pct = round(float(mr or 0)*100, 1) if mr else ""
            rows0.append({
                "계열": c["계열"], "지점": c["지점"], "훈련종류": c["훈련종류"],
                "과정명": c["과정명"],
                "시작일": fmt_mmdd(c["시작일"]) if c["시작일"] else "",
                "종료일": fmt_mmdd(c["종료일"]) if c["종료일"] else "",
                "정원": c["정원"],
                "확정인원": 확정 if 확정 != "" else "-",
                "모집률(%)": mr_pct if mr_pct != "" else "-",
                "개설상태": 개설라벨,
                "진행상태": c["진행상태"],
            })
        df0 = pd.DataFrame(rows0)

        def style_plan(row):
            s = row["개설상태"]
            if "미등록" in s and row["진행상태"] == "진행중":
                return ["background:#fff5f5"] * len(row)
            if "개강연기" in s or "폐강" in s:
                return ["background:#fffbeb"] * len(row)
            if "개강확정" in s:
                return ["background:#f0fff4"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df0.style.apply(style_plan, axis=1),
            use_container_width=True, hide_index=True,
        )


# ══════════════════════════════════════════════
# TAB MSG : 메신저 파싱
# ══════════════════════════════════════════════
with tab_msg:
    st.markdown("#### 📨 모집현황 입력")
    st.caption("메신저로 받은 보고 내용을 붙여넣으면 자동 파싱 → DB 저장 → 엑셀 다운로드까지 한번에 처리합니다.")

    # ── 기개강 과정 일괄 등록 ──────────────────────
    if recruit_bytes:
        with st.expander("📥 기개강 과정 일괄 등록 (모집현황 누적 파일 기반)", expanded=False):
            today_str = datetime.now().strftime("%Y-%m-%d")
            rec26 = parse_recruit_sheet(recruit_bytes, "26년")
            past_courses = [r for r in rec26 if r["시작일"] and r["시작일"] <= today_str]
            already = sum(1 for r in past_courses
                          if course_key(r["지점"], r["과정명"], r["회차"]) in db)
            st.caption(f"26년 시트 기개강 **{len(past_courses)}건** (기준일: {today_str}) — 이미 등록: {already}건")
            if past_courses:
                preview_df = pd.DataFrame([{
                    "월": r["월"], "계열": r["계열"], "지점": r["지점"],
                    "과정명": r["과정명"][:30], "시작일": r["시작일"],
                    "정원": r["정원"], "확정인원": r["확정인원"], "신청인원": r["신청인원"],
                    "모집률(%)": round(r["모집률"]*100, 1),
                    "신청률(%)": round(r["신청률"]*100, 1),
                    "등록여부": "✅" if course_key(r["지점"],r["과정명"],r["회차"]) in db else "❌",
                } for r in past_courses])
                st.dataframe(preview_df, use_container_width=True, hide_index=True)
                if st.button("💾 미등록 과정 일괄 등록 (개강확정)", type="primary", key="bulk_reg"):
                    imported = 0
                    for r in past_courses:
                        k = course_key(r["지점"], r["과정명"], r["회차"])
                        if k in db:
                            continue
                        record = {
                            "key": k, "계열": r["계열"], "지점": r["지점"],
                            "훈련종류": r["훈련종류"], "과정명": r["과정명"],
                            "시작일": r["시작일"], "종료일": r["종료일"],
                            "정원": r["정원"], "기준주차": r["월"],
                            "확정인원": r["확정인원"], "신청인원": r["신청인원"],
                            "모집률": r["모집률"], "신청률": r["신청률"],
                            "개설상태": "개강확정", "연기사유": "", "모집비고": "누적파일 자동등록",
                            "이수자평가예정":"","이수자평가신청일":"",
                            "평가완료":"","평가완료일":"","평가비고":"",
                            "비용단위기간":"","비용신청":"","비용금액":0,
                            "비용신청일":"","비용비고":"",
                            "취업_이수자":"","취업_취업자":"","취업_조사일":"","취업비고":"",
                            "만족도점수":"","만족도조사일":"","만족도비고":"",
                            "업데이트": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        }
                        save_record(k, record)
                        if not sheet: st.session_state.local_db[k] = record
                        db[k] = record
                        imported += 1
                    st.success(f"✅ {imported}개 과정 등록 완료!")
                    st.rerun()
    st.markdown("---")

    col_wk, col_info = st.columns([1, 2])
    with col_wk:
        week_label_m = st.text_input("기준 주차", placeholder="예: 3월 3주", key="msg_week")
    with col_info:
        if staff_map:
            st.info(f"👤 담당자 {len(staff_map)}명 로드됨 — 이름으로 계열/지점 자동 매핑", icon="✅")
        else:
            st.warning("사이드바에서 '지점 담당자 현황' 파일을 업로드하면 이름→지점 자동 매핑됩니다.", icon="⚠️")

    msg_text = st.text_area(
        "메신저 텍스트 붙여넣기 (한 명 또는 여러 명 한번에 가능)",
        placeholder="이정민\n안녕하세요! IT인천 모집현황 보고드립니다\n\n1. 과정명 : ...\n- 훈련기간 : 2026-03-20 ~ 2026-08-21\n- 모집인원 : 13명\n- 확정인원 : 13명\n\n고의정\n안녕하세요 컴퓨터종로 ...",
        height=280,
        key="msg_input",
    )

    btn_col, opt_col, reset_col = st.columns([2, 2, 1])
    with btn_col:
        do_parse = st.button("🔍 파싱 실행", type="primary", use_container_width=True)
    with opt_col:
        append_mode = st.checkbox(
            "📥 기존 결과에 추가 (순차 입력 시 체크)",
            value=False,
            key="msg_append",
            help="담당자별로 따로 붙여넣을 때 체크하면 결과가 누적됩니다."
        )
    with reset_col:
        if st.button("🗑 초기화", use_container_width=True):
            st.session_state.pop("parsed_results", None)
            st.rerun()

    if do_parse:
        if not msg_text.strip():
            st.error("텍스트를 입력해주세요.")
        else:
            with st.spinner("파싱 중..."):
                new_parsed = parse_messenger_all(msg_text.strip(), staff_map, courses)
            if append_mode and st.session_state.get("parsed_results"):
                # 중복 제거 (같은 보고자+과정명 이미 있으면 덮어쓰기)
                existing = st.session_state["parsed_results"]
                existing_keys = {(r["보고자"], r["과정명"]) for r in existing}
                merged = existing + [r for r in new_parsed if (r["보고자"], r["과정명"]) not in existing_keys]
                st.session_state["parsed_results"] = merged
                st.success(f"✅ {len(new_parsed)}개 추가 → 누적 {len(merged)}개")
            else:
                st.session_state["parsed_results"] = new_parsed
                if new_parsed:
                    st.success(f"✅ {len(new_parsed)}개 과정 파싱 완료!")
                else:
                    st.warning("과정을 찾지 못했습니다. 형식을 확인해주세요.")

    if st.session_state.get("parsed_results"):
        parsed = st.session_state["parsed_results"]
        st.markdown("---")

        # 요약 KPI
        kc = st.columns(4)
        reporters = list(dict.fromkeys(r["보고자"] for r in parsed if r["보고자"]))
        total_확정 = sum(r["확정인원"] for r in parsed)
        total_정원 = sum(r["정원"] for r in parsed)
        avg_mr = round(total_확정/total_정원*100, 1) if total_정원 > 0 else 0
        warn = sum(1 for r in parsed if r["정원"] > 0 and r["모집률(%)"] < 65)
        for col, num, lbl, clr in [
            (kc[0], len(reporters), f"보고 담당자 {','.join(reporters[:3])}{'...' if len(reporters)>3 else ''}", "#2b6cb0"),
            (kc[1], len(parsed),   "파싱 과정 수",   "#276749"),
            (kc[2], f"{avg_mr}%",  "평균 모집률",    "#e53e3e" if avg_mr < 65 else "#276749"),
            (kc[3], warn,          "모집 경고(65%↓)","#e53e3e"),
        ]:
            with col:
                st.markdown(
                    f'<div class="kpi-box"><div class="kpi-num" style="color:{clr};font-size:1.3rem">{num}</div>'
                    f'<div class="kpi-label">{lbl}</div></div>',
                    unsafe_allow_html=True,
                )
        st.markdown("<br>", unsafe_allow_html=True)

        # 결과 테이블 (직접 수정 가능)
        st.markdown("**📝 파싱 결과 — 셀을 클릭해 직접 수정 가능합니다**")
        edit_cols = ["보고자","계열","지점","훈련종류","과정명","시작일","종료일","훈련일수","훈련시간","정원","확정인원","신청인원","모집률(%)","신청률(%)","강의장","매칭과정명","비고"]
        df_edit = pd.DataFrame(parsed)[edit_cols].copy()
        # 타입 명시
        for _c in ["정원","확정인원","신청인원"]:
            df_edit[_c] = df_edit[_c].fillna(0).astype(int)
        for _c in ["모집률(%)","신청률(%)"]:
            df_edit[_c] = df_edit[_c].fillna(0.0).astype(float)
        for _c in ["보고자","계열","지점","훈련종류","과정명","시작일","종료일","훈련일수","훈련시간","강의장","매칭과정명","비고"]:
            df_edit[_c] = df_edit[_c].fillna("").astype(str)
        # 모집률/신청률은 재계산 표시용 (편집 불가)
        edited_df = st.data_editor(
            df_edit,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key="parsed_editor",
            column_config={
                "보고자":    st.column_config.TextColumn("보고자",    disabled=True, width="small"),
                "계열":      st.column_config.TextColumn("계열",      width="small"),
                "지점":      st.column_config.TextColumn("지점",      width="small"),
                "훈련종류":  st.column_config.TextColumn("훈련종류",  width="small"),
                "과정명":    st.column_config.TextColumn("과정명",    width="large"),
                "시작일":    st.column_config.TextColumn("시작일",    width="small"),
                "종료일":    st.column_config.TextColumn("종료일",    width="small"),
                "훈련일수":  st.column_config.TextColumn("훈련일수",  width="small"),
                "훈련시간":  st.column_config.TextColumn("훈련시간",  width="medium"),
                "정원":      st.column_config.NumberColumn("정원",    width="small", min_value=0, step=1),
                "확정인원":  st.column_config.NumberColumn("확정인원",width="small", min_value=0, step=1),
                "신청인원":  st.column_config.NumberColumn("신청인원",width="small", min_value=0, step=1),
                "모집률(%)": st.column_config.NumberColumn("모집률(%)", disabled=True, width="small"),
                "신청률(%)": st.column_config.NumberColumn("신청률(%)", disabled=True, width="small"),
                "강의장":    st.column_config.TextColumn("강의장",    width="medium"),
                "매칭과정명":st.column_config.TextColumn("매칭과정명",disabled=True, width="large"),
                "비고":      st.column_config.TextColumn("비고",      width="medium"),
            },
        )
        # 수정/삭제된 내용을 session_state에 반영 (모집률/신청률 재계산)
        new_parsed = []
        for _, row in edited_df.iterrows():
            if not str(row.get("과정명","")).strip():
                continue  # 과정명 없는 행(삭제된 행) 제외
            정원_e = int(row.get("정원", 0) or 0)
            확정_e = int(row.get("확정인원", 0) or 0)
            신청_e = int(row.get("신청인원", 0) or 0)
            new_parsed.append({
                "보고자":   str(row.get("보고자","")),
                "계열":     str(row.get("계열","")),
                "지점":     str(row.get("지점","")),
                "훈련종류": str(row.get("훈련종류","")),
                "과정명":   str(row.get("과정명","")),
                "시작일":   str(row.get("시작일","")),
                "종료일":   str(row.get("종료일","")),
                "훈련일수": str(row.get("훈련일수","")),
                "훈련시간": str(row.get("훈련시간","")),
                "정원":     정원_e,
                "확정인원": 확정_e,
                "신청인원": 신청_e,
                "모집률(%)": round(확정_e/정원_e*100, 1) if 정원_e > 0 else 0,
                "신청률(%)": round(신청_e/정원_e*100, 1) if 정원_e > 0 else 0,
                "강의장":   str(row.get("강의장","")),
                "매칭과정명": str(row.get("매칭과정명","")),
                "비고":     str(row.get("비고","")),
            })
        st.session_state["parsed_results"] = new_parsed
        parsed = new_parsed

        st.markdown("---")
        save_col, dl_col = st.columns(2)
        with save_col:
            if st.button("💾 DB에 저장 (모집현황 반영)", type="primary", use_container_width=True, key="msg_save"):
                if not week_label_m.strip():
                    st.error("기준 주차를 먼저 입력해주세요.")
                else:
                    saved_cnt = 0
                    for r in parsed:
                        정원 = int(r.get("정원") or 0)
                        확정 = int(r.get("확정인원") or 0)
                        신청 = int(r.get("신청인원") or 0)
                        k = course_key(r.get("지점",""), r.get("과정명",""), "1")
                        existing = db.get(k, {})
                        record = {
                            "key": k,
                            "계열": r.get("계열",""), "지점": r.get("지점",""),
                            "훈련종류": r.get("훈련종류",""), "과정명": r.get("과정명",""),
                            "시작일": r.get("시작일",""), "종료일": r.get("종료일",""),
                            "정원": 정원,
                            "기준주차": week_label_m.strip(),
                            "확정인원": 확정, "신청인원": 신청,
                            "모집률": round(확정/정원, 4) if 정원 > 0 else 0,
                            "신청률": round(신청/정원, 4) if 정원 > 0 else 0,
                            "개설상태": "개강확정",
                            "연기사유": "", "모집비고": r.get("비고",""),
                            "이수자평가예정": existing.get("이수자평가예정",""),
                            "이수자평가신청일": existing.get("이수자평가신청일",""),
                            "평가완료": existing.get("평가완료",""),
                            "평가완료일": existing.get("평가완료일",""),
                            "평가비고": existing.get("평가비고",""),
                            "비용단위기간": existing.get("비용단위기간",""),
                            "비용신청": existing.get("비용신청",""),
                            "비용금액": existing.get("비용금액",0),
                            "비용신청일": existing.get("비용신청일",""),
                            "비용비고": existing.get("비용비고",""),
                            "취업_이수자": existing.get("취업_이수자",""),
                            "취업_취업자": existing.get("취업_취업자",""),
                            "취업_조사일": existing.get("취업_조사일",""),
                            "취업비고": existing.get("취업비고",""),
                            "만족도점수": existing.get("만족도점수",""),
                            "만족도조사일": existing.get("만족도조사일",""),
                            "만족도비고": existing.get("만족도비고",""),
                            "업데이트": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        }
                        save_record(k, record)
                        if not sheet: st.session_state.local_db[k] = record
                        db[k] = record
                        saved_cnt += 1
                    st.success(f"✅ {saved_cnt}개 과정 DB 저장 완료! ({week_label_m})")
                    st.rerun()
        with dl_col:
            if week_label_m.strip():
                excel_bytes = export_messenger_excel(parsed, week_label_m.strip())
                st.download_button(
                    label=f"📥 엑셀 다운로드",
                    data=excel_bytes,
                    file_name=f"(개설과정 조사) {week_label_m}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            else:
                st.info("기준 주차 입력 시 엑셀 다운로드 활성화", icon="💡")

# ══════════════════════════════════════════════
# TAB 2 : 모집현황 조회
# ══════════════════════════════════════════════
with tab2:
    recs = [v for v in db.values() if v.get("개설상태","미입력") not in ("","미입력")]

    if not recs:
        st.info("저장된 모집현황이 없습니다. '모집현황 입력' 탭에서 먼저 데이터를 입력하세요.")
    else:
        # 월별 필터 (훈련시작일 기준)
        def get_month(r):
            s = str(r.get("시작일",""))
            return s[:7] if len(s) >= 7 else "미정"
        month_set = sorted(set(get_month(r) for r in recs))
        month_labels = {m: (f"{int(m[5:7])}월" if len(m) >= 7 and m[5:7].isdigit() else m) for m in month_set}
        sel_month = st.radio(
            "훈련시작 월", ["전체"] + month_set,
            format_func=lambda x: "전체" if x == "전체" else month_labels.get(x, x),
            horizontal=True,
        )
        recs = recs if sel_month == "전체" else [r for r in recs if get_month(r) == sel_month]

        fc1, fc2, fc3, fc4 = st.columns(4)
        with fc1: sf2 = st.multiselect("계열", sorted(set(r.get("계열","") for r in recs)))
        with fc2: bf2 = st.multiselect("지점", sorted(set(r.get("지점","") for r in recs)))
        with fc3: stf = st.multiselect("개설상태", ["준비중","개강확정","개강연기","폐강"],
                                        default=["준비중","개강확정","개강연기","폐강"])
        with fc4: warn_only = st.checkbox("경고 과정만")

        filtered = [
            r for r in recs
            if (not sf2 or r.get("계열") in sf2)
            and (not bf2 or r.get("지점") in bf2)
            and r.get("개설상태") in stf
        ]
        if warn_only:
            filtered = [
                r for r in filtered
                if r.get("개설상태") == "개강확정"
                and (float(r.get("모집률",1) or 1) < 0.65
                     or float(r.get("신청률",1) or 1) < 0.70)
            ]

        # 요약 KPI
        개설 = [r for r in filtered if r.get("개설상태") == "개강확정"]
        경고 = [r for r in 개설
                if float(r.get("모집률",1) or 1) < 0.65
                or float(r.get("신청률",1) or 1) < 0.70]
        연기 = [r for r in filtered if r.get("개설상태") == "개강연기"]
        폐강 = [r for r in filtered if r.get("개설상태") == "폐강"]

        mc = st.columns(5)
        for col, num, lbl, clr in [
            (mc[0], len(filtered), "전체",    "#2b6cb0"),
            (mc[1], len(개설),     "개강확정", "#276749"),
            (mc[2], len(경고),     "모집경고", "#e53e3e"),
            (mc[3], len(연기),     "개강연기", "#c05621"),
            (mc[4], len(폐강),     "폐강",     "#718096"),
        ]:
            with col:
                st.markdown(
                    f'<div class="kpi-box"><div class="kpi-num" style="color:{clr};font-size:1.4rem">{num}</div>'
                    f'<div class="kpi-label">{lbl}</div></div>',
                    unsafe_allow_html=True,
                )
        st.markdown("<br>", unsafe_allow_html=True)

        # 테이블 (pandas styler)
        rows = []
        for r in filtered:
            mp = float(r.get("모집률", 0) or 0) * 100
            sp = float(r.get("신청률", 0) or 0) * 100
            rows.append({
                "계열":      r.get("계열",""),
                "지점":      r.get("지점",""),
                "훈련종류":  r.get("훈련종류",""),
                "과정명":    r.get("과정명",""),
                "기준주차":  r.get("기준주차",""),
                "정원":      int(r.get("정원",0) or 0),
                "확정":      int(r.get("확정인원",0) or 0),
                "신청":      int(r.get("신청인원",0) or 0),
                "모집률(%)": round(mp, 1),
                "신청률(%)": round(sp, 1),
                "개설상태":  r.get("개설상태",""),
                "연기사유":  r.get("연기사유",""),
                "비고":      r.get("모집비고",""),
            })

        df = pd.DataFrame(rows)

        def style_row(row):
            s = row["개설상태"]
            if s == "준비중":
                return ["color:#718096;background:#f0f4f8"] * len(row)
            if s == "폐강":
                return ["color:#aaa;background:#f7fafc"] * len(row)
            if s == "개강연기":
                return ["background:#fffbeb"] * len(row)
            if s == "개강확정" and (row["모집률(%)"] < 65 or row["신청률(%)"] < 70):
                return ["background:#fff5f5;color:#c53030"] * len(row)
            return [""] * len(row)

        styled = (
            df.style
            .apply(style_row, axis=1)
            .map(lambda v: "color:#e53e3e;font-weight:700" if v < 65 else
                           "color:#276749;font-weight:700", subset=["모집률(%)"])
            .map(lambda v: "color:#e53e3e;font-weight:700" if v < 70 else
                           "color:#276749;font-weight:700", subset=["신청률(%)"])
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # 경고 과정 하이라이트
        if 경고:
            st.markdown(f"**⚠️ 모집경고 과정 {len(경고)}건**")
            for r in 경고:
                mp = float(r.get("모집률",0) or 0) * 100
                sp = float(r.get("신청률",0) or 0) * 100
                st.error(
                    f"**{r.get('과정명','')}** ({r.get('지점','')})  |  "
                    f"확정 {r.get('확정인원',0)}/{r.get('정원',0)}명  |  "
                    f"모집률 **{mp:.1f}%**  신청률 **{sp:.1f}%**"
                    + (f"  |  비고: {r.get('모집비고','')}" if r.get("모집비고") else "")
                )

# ══════════════════════════════════════════════
# TAB 3 : 과정 추적 관리
# ══════════════════════════════════════════════
with tab3:
    st.markdown("#### 개강확정 과정별 성과 관리")

    confirmed_keys = [k for k, v in db.items() if v.get("개설상태") == "개강확정"]

    if not confirmed_keys:
        st.info("개강확정 과정이 없습니다. 모집현황 입력 탭에서 데이터를 저장하면 표시됩니다.")
    else:
        br_t = st.multiselect(
            "지점 필터",
            sorted(set(db[k].get("지점","") for k in confirmed_keys)),
            key="track_br",
        )
        keys_show = [k for k in confirmed_keys
                     if not br_t or db[k].get("지점") in br_t]

        # ── 4개 성과 탭 ─────────────────────────────
        t3a, t3b, t3c, t3d = st.tabs(["📝 이수자평가 현황", "💰 비용신청 현황", "💼 취업 현황", "⭐ 만족도 현황"])

        def course_label(r):
            mp = float(r.get("모집률",0) or 0) * 100
            return f"**{r.get('과정명','')}** — {r.get('지점','')}  |  {str(r.get('시작일',''))[:7]}  |  모집률 {mp:.1f}%"

        # ── 이수자평가 현황 ──────────────────────────
        with t3a:
            st.caption("이수자 평가 신청 및 결과를 관리합니다.")
            for key in keys_show:
                r = db[key]
                ev_planned = str(r.get("이수자평가예정","")) == "True"
                ev_done    = str(r.get("평가완료","")) == "True"
                badge = "✅완료" if ev_done else ("📋예정" if ev_planned else "⬜미등록")
                with st.expander(f"{badge}  {course_label(r)}", expanded=False):
                    with st.form(key=f"eva_{key}"):
                        a1, a2, a3 = st.columns(3)
                        with a1:
                            ev_pl = st.checkbox("평가 예정", value=ev_planned, key=f"evpa_{key}")
                            ev_yn = st.checkbox("평가 완료", value=ev_done, key=f"evya_{key}")
                        with a2:
                            try:
                                evd_def = datetime.strptime(r["이수자평가신청일"],"%Y-%m-%d").date() \
                                    if r.get("이수자평가신청일") else datetime.today().date()
                            except Exception:
                                evd_def = datetime.today().date()
                            ev_신청일 = st.date_input("신청 예정일", value=evd_def, key=f"evsd_{key}")
                            try:
                                evc_def = datetime.strptime(r["평가완료일"],"%Y-%m-%d").date() \
                                    if r.get("평가완료일") else datetime.today().date()
                            except Exception:
                                evc_def = datetime.today().date()
                            ev_완료일 = st.date_input("완료일", value=evc_def, key=f"evcd_{key}")
                        with a3:
                            ev_nt = st.text_input("결과/비고", value=r.get("평가비고",""), key=f"evna_{key}",
                                                  placeholder="합격/불합격/진행중")
                        if st.form_submit_button("💾 저장", use_container_width=True):
                            updated = dict(r)
                            updated.update({
                                "이수자평가예정":   str(ev_pl),
                                "이수자평가신청일": ev_신청일.strftime("%Y-%m-%d"),
                                "평가완료":  str(ev_yn),
                                "평가완료일": ev_완료일.strftime("%Y-%m-%d"),
                                "평가비고": ev_nt,
                                "업데이트": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            })
                            save_record(key, updated)
                            if not sheet: st.session_state.local_db[key] = updated
                            db[key] = updated
                            st.success("저장 완료!"); st.rerun()

        # ── 비용신청 현황 ────────────────────────────
        with t3b:
            st.caption("단위기간별 훈련비용 신청 현황을 관리합니다.")
            for key in keys_show:
                r = db[key]
                기존_단위 = parse_비용단위기간(r.get("비용단위기간",""))
                단위기간_list = calc_단위기간(r.get("시작일",""), r.get("종료일",""))
                완료수 = sum(1 for v in 기존_단위.values() if v["완료"])
                전체수 = len(단위기간_list)
                badge = f"✅{완료수}/{전체수}회차" if 완료수 > 0 else "⬜미신청"
                with st.expander(f"{badge}  {course_label(r)}", expanded=False):
                    if not 단위기간_list:
                        st.warning("시작일/종료일 정보가 없어 단위기간을 계산할 수 없습니다.")
                    else:
                        st.caption(f"훈련기간: {r.get('시작일','')} ~ {r.get('종료일','')}  |  총 {전체수}회차")
                        with st.form(key=f"csb_{key}"):
                            new_단위 = {}
                            for p in 단위기간_list:
                                k_p = p["key"]
                                prev = 기존_단위.get(k_p, {"완료": False, "금액": 0})
                                pc1, pc2, pc3 = st.columns([2, 1, 2])
                                with pc1:
                                    st.markdown(
                                        f"**{p['회차']}회차** &nbsp; `{p['시작']} ~ {p['종료']}`"
                                    )
                                with pc2:
                                    yn = st.checkbox("신청완료", value=prev["완료"],
                                                     key=f"csyn_{key}_{k_p}")
                                with pc3:
                                    am = st.number_input("금액(원)", value=int(prev["금액"]),
                                                         min_value=0, step=10000,
                                                         key=f"csam_{key}_{k_p}")
                                new_단위[k_p] = {"완료": yn, "금액": am}
                            cs_nt = st.text_input("비고", value=r.get("비용비고",""),
                                                  key=f"csnb_{key}")
                            if st.form_submit_button("💾 저장", use_container_width=True):
                                total_am = sum(v["금액"] for v in new_단위.values() if v["완료"])
                                all_done = all(v["완료"] for v in new_단위.values())
                                updated = dict(r)
                                updated.update({
                                    "비용단위기간": serialize_비용단위기간(new_단위),
                                    "비용신청":  str(all_done),
                                    "비용금액":  total_am,
                                    "비용비고":  cs_nt,
                                    "업데이트":  datetime.now().strftime("%Y-%m-%d %H:%M"),
                                })
                                save_record(key, updated)
                                if not sheet: st.session_state.local_db[key] = updated
                                db[key] = updated
                                st.success("저장 완료!"); st.rerun()

        # ── 취업 현황 ────────────────────────────────
        with t3c:
            st.caption("이수자 취업 성과를 관리합니다.")
            for key in keys_show:
                r = db[key]
                em_tot = int(r.get("취업_이수자",0) or 0)
                em_hi  = int(r.get("취업_취업자",0) or 0)
                emp_rate = f"{em_hi/em_tot*100:.1f}%" if em_tot > 0 else "미입력"
                badge = f"💼{emp_rate}" if em_tot > 0 else "⬜미입력"
                with st.expander(f"{badge}  {course_label(r)}", expanded=False):
                    with st.form(key=f"emc_{key}"):
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            em_total = st.number_input("이수자수", value=em_tot, min_value=0, key=f"emtc_{key}")
                            em_hired = st.number_input("취업자수", value=em_hi,  min_value=0, key=f"emhc_{key}")
                        with c2:
                            try:
                                emd_def = datetime.strptime(r["취업_조사일"],"%Y-%m-%d").date() \
                                    if r.get("취업_조사일") else datetime.today().date()
                            except Exception:
                                emd_def = datetime.today().date()
                            em_dt = st.date_input("조사일", value=emd_def, key=f"emdc_{key}")
                            if em_total > 0:
                                st.metric("취업률", f"{em_hired/em_total*100:.1f}%")
                        with c3:
                            em_nt = st.text_input("비고", value=r.get("취업비고",""), key=f"emnc_{key}")
                        if st.form_submit_button("💾 저장", use_container_width=True):
                            updated = dict(r)
                            updated.update({
                                "취업_이수자": em_total, "취업_취업자": em_hired,
                                "취업_조사일": em_dt.strftime("%Y-%m-%d"), "취업비고": em_nt,
                                "업데이트": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            })
                            save_record(key, updated)
                            if not sheet: st.session_state.local_db[key] = updated
                            db[key] = updated
                            st.success("저장 완료!"); st.rerun()

        # ── 만족도 현황 ──────────────────────────────
        with t3d:
            st.caption("훈련 만족도 조사 결과를 관리합니다.")
            for key in keys_show:
                r = db[key]
                sat = float(r.get("만족도점수",0) or 0)
                badge = f"⭐{sat:.1f}" if sat > 0 else "⬜미입력"
                with st.expander(f"{badge}  {course_label(r)}", expanded=False):
                    with st.form(key=f"satd_{key}"):
                        d1, d2, d3 = st.columns(3)
                        with d1:
                            sat_sc = st.number_input("점수 (0~5)", value=sat,
                                                     min_value=0.0, max_value=5.0, step=0.1,
                                                     key=f"sasd_{key}")
                        with d2:
                            try:
                                sadd_def = datetime.strptime(r["만족도조사일"],"%Y-%m-%d").date() \
                                    if r.get("만족도조사일") else datetime.today().date()
                            except Exception:
                                sadd_def = datetime.today().date()
                            sat_dt = st.date_input("조사일", value=sadd_def, key=f"sadd_{key}")
                        with d3:
                            sat_nt = st.text_input("비고", value=r.get("만족도비고",""), key=f"sand_{key}")
                        if st.form_submit_button("💾 저장", use_container_width=True):
                            updated = dict(r)
                            updated.update({
                                "만족도점수": sat_sc,
                                "만족도조사일": sat_dt.strftime("%Y-%m-%d"), "만족도비고": sat_nt,
                                "업데이트": datetime.now().strftime("%Y-%m-%d %H:%M"),
                            })
                            save_record(key, updated)
                            if not sheet: st.session_state.local_db[key] = updated
                            db[key] = updated
                            st.success("저장 완료!"); st.rerun()

        # ── 집계 ─────────────────────────────────────
        st.markdown("---")
        st.markdown("**📊 성과 집계**")
        agg = st.columns(4)
        ev_done_cnt  = sum(1 for k in confirmed_keys if str(db[k].get("평가완료","")) == "True")
        cs_done_cnt  = sum(1 for k in confirmed_keys if str(db[k].get("비용신청","")) == "True")
        em_list  = [(int(db[k].get("취업_이수자",0) or 0), int(db[k].get("취업_취업자",0) or 0))
                    for k in confirmed_keys if int(db[k].get("취업_이수자",0) or 0) > 0]
        avg_emp  = (sum(h/t for t,h in em_list)/len(em_list)*100) if em_list else 0
        sat_list = [float(db[k].get("만족도점수",0) or 0)
                    for k in confirmed_keys if float(db[k].get("만족도점수",0) or 0) > 0]
        avg_sat  = sum(sat_list)/len(sat_list) if sat_list else 0

        for col, num, lbl, clr in [
            (agg[0], f"{ev_done_cnt}/{len(confirmed_keys)}", "이수자평가 완료", "#2b6cb0"),
            (agg[1], f"{cs_done_cnt}/{len(confirmed_keys)}", "비용신청 완료",   "#276749"),
            (agg[2], f"{avg_emp:.1f}%",                      "평균 취업률",     "#553c9a"),
            (agg[3], f"{avg_sat:.2f}/5.0",                   "평균 만족도",     "#744210"),
        ]:
            with col:
                st.markdown(
                    f'<div class="kpi-box"><div class="kpi-num" style="color:{clr};font-size:1.35rem">{num}</div>'
                    f'<div class="kpi-label">{lbl}</div></div>',
                    unsafe_allow_html=True,
                )

# ══════════════════════════════════════════════
# TAB 4 : 반납 분석
# ══════════════════════════════════════════════
with tab4:
    from collections import Counter
    rc = Counter(r["사유분류"] for r in returns)

    if not rc:
        st.info("반납 과정이 없습니다.")
    else:
        COLORS = {
            "강의장 부족":"#FED7D7","모집률 저조":"#FEFCBF","미개설":"#EDF2F7",
            "강사 문제":"#BEE3F8","비용/효율성":"#C6F6D5","직종 중복/조정":"#E9D8FD",
            "내부 결정":"#FBD38D","기타":"#EDF2F7",
        }
        TC = {
            "강의장 부족":"#822727","모집률 저조":"#744210","미개설":"#4a5568",
            "강사 문제":"#2a4365","비용/효율성":"#276749","직종 중복/조정":"#553c9a",
            "내부 결정":"#7b341e","기타":"#4a5568",
        }
        rcols = st.columns(min(len(rc), 4))
        for i, (reason, cnt) in enumerate(sorted(rc.items(), key=lambda x: -x[1])):
            bg = COLORS.get(reason,"#EDF2F7"); tc2 = TC.get(reason,"#4a5568")
            with rcols[i % 4]:
                st.markdown(
                    f"<div style='background:{bg};border-radius:8px;padding:0.8rem;"
                    f"text-align:center;margin-bottom:0.5rem'>"
                    f"<div style='font-size:1.5rem;font-weight:700;color:{tc2}'>{cnt}건</div>"
                    f"<div style='font-size:0.8rem;color:{tc2};font-weight:600'>{reason}</div></div>",
                    unsafe_allow_html=True,
                )

        sel = st.selectbox("사유 필터", ["전체"] + list(rc.keys()))
        show = returns if sel == "전체" else [r for r in returns if r["사유분류"] == sel]
        st.dataframe(
            pd.DataFrame([{
                "계열":r["계열"],"지점":r["지점"],"훈련종류":r["훈련종류"],
                "과정명":r["과정명"],"사유분류":r["사유분류"],"반납사유":r["반납사유"],
            } for r in show]),
            use_container_width=True, hide_index=True,
        )

# ══════════════════════════════════════════════
# TAB 5 : 인증평가 현황
# ══════════════════════════════════════════════
@st.cache_data(show_spinner="인증평가 파일 읽는 중...")
def parse_cert(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    cur_series = ""
    for row in ws.iter_rows(min_row=7, values_only=True):
        if not row[2]:   # 지점명 없으면 skip
            continue
        if row[1]:
            cur_series = str(row[1]).strip()
        def v(x): return x if x is not None else ""
        def f(x):
            try: return round(float(x), 2)
            except: return ""
        # 유효기간 파싱 (S열=index 18): "2024.01.01~2026.12.31"
        유효기간_raw = str(v(row[18])).strip()
        end_year = ""
        if "~" in 유효기간_raw:
            try:
                end_part = 유효기간_raw.split("~")[1].strip()
                end_year = int(end_part[:4])
            except: pass
        # 등급 파싱 + 다음 평가 연도 계산
        등급_raw = str(v(row[3])).strip()
        if "5년" in 등급_raw:    등급분류 = "5년 우수"
        elif "3년" in 등급_raw:  등급분류 = "3년 인증"
        elif "1년" in 등급_raw:  등급분류 = "1년 인증"
        elif "유예" in 등급_raw: 등급분류 = "인증유예"
        else:                     등급분류 = 등급_raw
        # 다음 평가 연도: 등급(YY) + 기간 → ex) 3년인증(25) → 2025+3=2028
        next_eval_year = None
        m_grade = re.search(r"(\d+)년\s*(?:인증|우수)\((\d{2})\)", 등급_raw)
        THIS_YEAR = datetime.now().year
        if m_grade:
            period     = int(m_grade.group(1))
            start_yy   = int(m_grade.group(2))
            start_full = 2000 + start_yy
            # 유효기간 마지막 해에 평가: 3년 인증(24) → 2026년 평가
            next_eval_year = start_full + period - 1
            # 올해 새로 받은 인증은 제외 (1년/3년/5년 모두)
            if start_full == THIS_YEAR:
                next_eval_year = None
        # 유효기간 컬럼이 명시된 경우: 그 기간 안에는 평가 대상 아님
        # 유효기간 끝나는 해에 평가 (end_year가 있고 미래라면 그 해가 eval)
        if end_year and isinstance(end_year, int):
            if end_year > THIS_YEAR:
                # 유효기간 남아있음 → 평가 대상 아님 (end_year 해에 평가 예정)
                next_eval_year = end_year
            elif end_year == THIS_YEAR:
                # 유효기간이 올해 끝남 → 올해 평가 대상
                next_eval_year = THIS_YEAR
            # end_year < THIS_YEAR: 유효기간 이미 지남 → grade 기반 eval_year 유지
        올해대상 = (next_eval_year == THIS_YEAR) or (등급분류 == "인증유예")
        내년대상 = (next_eval_year == THIS_YEAR + 1) if next_eval_year else False
        rows.append({
            "계열": cur_series,
            "지점": str(v(row[2])).strip(),
            "평가등급": 등급_raw,
            "등급분류": 등급분류,
            "역량평가총점": f(row[4]),
            "훈련성과총점": f(row[5]),
            "실업점수":     f(row[6]),
            "근로점수":     f(row[7]),
            "일반취업률":   f(row[8]),
            "고용유지":     f(row[9]),
            "만족도_실업자":f(row[10]),
            "만족도_재직자":f(row[11]),
            "훈련비중_실업":str(v(row[12])),
            "훈련비중_근로":str(v(row[13])),
            "현장평가총점": f(row[14]),
            "과정관리":     f(row[15]),
            "인프라":       f(row[16]),
            "전담인력":     f(row[17]),
            "유효기간":      유효기간_raw,
            "만료년도":      end_year,
            "다음평가연도":  next_eval_year,
            "올해평가대상":  올해대상,
            "내년평가대상":  내년대상,
        })
    return rows

with tab5:
    st.markdown("#### 🏅 인증평가 현황")

    AUTO_CERT = "cert.xlsx.xlsx"
    with st.sidebar:
        st.markdown("### 🏅 인증평가 파일")
        if os.path.exists(AUTO_CERT):
            st.success("✅ cert.xlsx 자동 로드됨", icon="🏅")
            cert_file = AUTO_CERT
            cert_override = st.file_uploader("인증평가 파일 교체", type=["xlsx","XLSX"],
                                              key="cert_up")
            if cert_override: cert_file = cert_override
        else:
            cert_file = st.file_uploader("인증평가 엑셀 업로드", type=["xlsx","XLSX"],
                                          key="cert_up")

    if not cert_file:
        st.info("👈 왼쪽 사이드바에서 인증평가 엑셀 파일을 업로드해주세요.")
    else:
        if isinstance(cert_file, str):
            with open(cert_file,"rb") as f: cert_bytes = f.read()
        else:
            cert_bytes = cert_file.read()
        cert_rows = parse_cert(cert_bytes)

        # 등급별 스타일
        GRADE_STYLE = {
            "5년 우수":  ("🥇","#744210","#FEFCBF"),
            "3년 인증":  ("✅","#276749","#C6F6D5"),
            "1년 인증":  ("⚠️","#c05621","#FEEBC8"),
            "인증유예":  ("🔴","#822727","#FED7D7"),
        }

        # KPI
        THIS_YEAR = datetime.now().year
        total_c = len(cert_rows)
        g5 = sum(1 for r in cert_rows if r["등급분류"]=="5년 우수")
        g3 = sum(1 for r in cert_rows if r["등급분류"]=="3년 인증")
        g1 = sum(1 for r in cert_rows if r["등급분류"]=="1년 인증")
        gy = sum(1 for r in cert_rows if r["등급분류"]=="인증유예")
        target_올해 = sum(1 for r in cert_rows if r["올해평가대상"])
        target_내년 = sum(1 for r in cert_rows if r["내년평가대상"])

        kc = st.columns(6)
        for col, num, lbl, clr in [
            (kc[0], total_c,      "전체 지점",         "#2b6cb0"),
            (kc[1], g5,           "5년 우수",          "#744210"),
            (kc[2], g3,           "3년 인증",          "#276749"),
            (kc[3], g1+gy,        "1년/유예",          "#c05621"),
            (kc[4], target_올해,  f"{THIS_YEAR}년 대상","#e53e3e"),
            (kc[5], target_내년,  f"{THIS_YEAR+1}년 대상","#c05621"),
        ]:
            with col:
                st.markdown(
                    f'<div class="kpi-box"><div class="kpi-num" style="color:{clr}">{num}</div>'
                    f'<div class="kpi-label">{lbl}</div></div>',
                    unsafe_allow_html=True,
                )
        st.markdown("<br>", unsafe_allow_html=True)

        # 필터
        cf1, cf2, cf3 = st.columns(3)
        with cf1:
            sel_ser  = st.multiselect("계열", sorted(set(r["계열"] for r in cert_rows)), key="cf_ser")
        with cf2:
            sel_grade= st.multiselect("등급", ["5년 우수","3년 인증","1년 인증","인증유예"], key="cf_gr")
        with cf3:
            target_filter = st.radio("평가대상 필터", ["전체", f"{THIS_YEAR}년 대상", f"{THIS_YEAR+1}년 대상"],
                                     horizontal=True, key="cf_target")

        filtered_c = [
            r for r in cert_rows
            if (not sel_ser   or r["계열"]    in sel_ser)
            and (not sel_grade or r["등급분류"] in sel_grade)
            and (target_filter == "전체"
                 or (target_filter == f"{THIS_YEAR}년 대상" and r["올해평가대상"])
                 or (target_filter == f"{THIS_YEAR+1}년 대상" and r["내년평가대상"]))
        ]

        # 올해/내년 평가대상 경고
        target_list = [r for r in filtered_c if r["올해평가대상"]]
        next_list   = [r for r in filtered_c if r["내년평가대상"]]
        if target_list:
            st.error(f"🔴 **{THIS_YEAR}년 평가대상 {len(target_list)}개 지점** — 올해 평가 필요")
        if next_list:
            st.warning(f"⚠️ **{THIS_YEAR+1}년 평가대상 {len(next_list)}개 지점** — 내년 평가 예정")

        # 테이블
        disp_rows = []
        for r in filtered_c:
            ic, fc, bc = GRADE_STYLE.get(r["등급분류"], ("","#000","#fff"))
            disp_rows.append({
                "계열":         r["계열"],
                "지점":         r["지점"],
                "평가등급":     r["평가등급"],
                "다음평가":     str(r["다음평가연도"]) if r["다음평가연도"] else "",
                "평가대상":     "🔴 올해" if r["올해평가대상"] else ("⚠️ 내년" if r["내년평가대상"] else ""),
                "역량평가총점": r["역량평가총점"],
                "훈련성과총점": r["훈련성과총점"],
                "실업점수":     r["실업점수"],
                "근로점수":     r["근로점수"],
                "일반취업률":   r["일반취업률"],
                "고용유지":     r["고용유지"],
                "만족도_실업":  r["만족도_실업자"],
                "만족도_재직":  r["만족도_재직자"],
                "현장평가총점": r["현장평가총점"],
                "과정관리":     r["과정관리"],
                "인프라":       r["인프라"],
                "전담인력":     r["전담인력"],
                "유효기간":     r["유효기간"],
            })

        df_c = pd.DataFrame(disp_rows)

        def style_cert(row):
            등급 = ""
            for r in filtered_c:
                if r["지점"] == row["지점"] and r["계열"] == row["계열"]:
                    등급 = r["등급분류"]; break
            if 등급 == "5년 우수": return ["background:#FEFCBF"] * len(row)
            if 등급 == "1년 인증": return ["background:#FEEBC8"] * len(row)
            if 등급 == "인증유예": return ["background:#FED7D7"] * len(row)
            return [""] * len(row)

        st.dataframe(
            df_c.style.apply(style_cert, axis=1),
            use_container_width=True, hide_index=True,
        )

        # 계열별 평균 요약
        st.markdown("---")
        st.markdown("**📊 계열별 평균 역량평가 총점**")
        from collections import defaultdict
        ser_scores = defaultdict(list)
        for r in cert_rows:
            if isinstance(r["역량평가총점"], float):
                ser_scores[r["계열"]].append(r["역량평가총점"])
        sum_cols = st.columns(len(ser_scores) or 1)
        for i, (ser, scores) in enumerate(sorted(ser_scores.items())):
            avg = sum(scores)/len(scores)
            with sum_cols[i % len(sum_cols)]:
                st.markdown(
                    f'<div class="kpi-box"><div class="kpi-num" style="color:#2b6cb0;font-size:1.3rem">{avg:.1f}</div>'
                    f'<div class="kpi-label">{ser}</div></div>',
                    unsafe_allow_html=True,
                )

# ══════════════════════════════════════════════
# TAB 6 : 연도별 비교
# ══════════════════════════════════════════════
with tab6:
    st.markdown("#### 📈 연도별 모집현황 비교")

    if not recruit_bytes:
        st.info("👈 사이드바에서 '모집현황 누적 파일'을 업로드하면 연도별 비교를 볼 수 있습니다.")
    else:
        CUR_YEAR = datetime.now().year
        CMP_YEARS = [str(CUR_YEAR), str(CUR_YEAR-1), str(CUR_YEAR-2)]

        # 각 연도 시트 파싱
        year_data = {}
        for yr in CMP_YEARS:
            sheet_name = f"{yr[2:]}년"  # "26년", "25년", "24년"
            rows = parse_recruit_sheet(recruit_bytes, sheet_name)
            if rows:
                year_data[yr] = rows

        # 누적 시트 요약
        summary = parse_recruit_summary(recruit_bytes)

        # ── 필터 ─────────────────────────────────────
        fc1, fc2 = st.columns(2)
        with fc1:
            all_series = sorted(set(r["계열"] for rows in year_data.values() for r in rows if r["계열"]))
            sel_ser_c = st.multiselect("계열 필터", all_series, key="cmp_ser")
        with fc2:
            sel_month_c = st.multiselect("월 필터", ["1월","2월","3월","4월","5월","6월",
                                                      "7월","8월","9월","10월","11월","12월"], key="cmp_mon")

        def filter_rows(rows):
            return [r for r in rows
                    if (not sel_ser_c or r["계열"] in sel_ser_c)
                    and (not sel_month_c or r["월"] in sel_month_c)]

        # ── 연도별 KPI 비교 ───────────────────────────
        st.markdown("---")
        st.markdown("**전체 집계 비교**")
        kpi_cols = st.columns(len(year_data))
        for ci, (yr, rows) in enumerate(sorted(year_data.items(), reverse=True)):
            fr = filter_rows(rows)
            tot_정원   = sum(r["정원"] for r in fr)
            tot_확정   = sum(r["확정인원"] for r in fr)
            tot_신청   = sum(r["신청인원"] for r in fr)
            avg_모집   = tot_확정/tot_정원*100 if tot_정원 > 0 else 0
            avg_신청   = tot_신청/tot_정원*100 if tot_정원 > 0 else 0
            with kpi_cols[ci]:
                st.markdown(
                    f"<div class='kpi-box'>"
                    f"<div class='kpi-num' style='color:#2b6cb0;font-size:1.4rem'>{yr}년</div>"
                    f"<div style='font-size:0.85rem;margin-top:0.3rem'>"
                    f"과정수 <b>{len(fr)}</b>건<br>"
                    f"정원 <b>{tot_정원}</b>명 / 확정 <b>{tot_확정}</b>명<br>"
                    f"모집률 <b style='color:{'#e53e3e' if avg_모집<65 else '#276749'}'>{avg_모집:.1f}%</b> &nbsp; "
                    f"신청률 <b style='color:{'#e53e3e' if avg_신청<70 else '#276749'}'>{avg_신청:.1f}%</b>"
                    f"</div></div>",
                    unsafe_allow_html=True
                )

        # ── 월별 비교 테이블 ──────────────────────────
        st.markdown("---")
        st.markdown("**월별 모집률 비교**")
        MONTHS = ["1월","2월","3월","4월","5월","6월","7월","8월","9월","10월","11월","12월"]
        comp_rows = []
        for mon in MONTHS:
            row_d = {"월": mon}
            for yr, rows in sorted(year_data.items(), reverse=True):
                fr = [r for r in filter_rows(rows) if r["월"] == mon]
                if fr:
                    tot_정 = sum(r["정원"] for r in fr)
                    tot_확 = sum(r["확정인원"] for r in fr)
                    tot_신 = sum(r["신청인원"] for r in fr)
                    row_d[f"{yr}년 과정수"]   = len(fr)
                    row_d[f"{yr}년 모집률(%)"] = round(tot_확/tot_정*100,1) if tot_정>0 else 0
                    row_d[f"{yr}년 신청률(%)"] = round(tot_신/tot_정*100,1) if tot_정>0 else 0
                else:
                    row_d[f"{yr}년 과정수"]    = 0
                    row_d[f"{yr}년 모집률(%)"] = 0
                    row_d[f"{yr}년 신청률(%)"] = 0
            comp_rows.append(row_d)

        df_comp = pd.DataFrame([r for r in comp_rows if any(r.get(f"{yr}년 과정수",0)>0 for yr in year_data)])

        def style_comp(val):
            if isinstance(val, float):
                if val > 0 and val < 65: return "color:#e53e3e;font-weight:700"
                if val >= 65: return "color:#276749;font-weight:700"
            return ""

        mr_cols = [c for c in df_comp.columns if "모집률" in c or "신청률" in c]
        st.dataframe(
            df_comp.style.map(style_comp, subset=mr_cols),
            use_container_width=True, hide_index=True
        )

        # ── 계열별 비교 ───────────────────────────────
        st.markdown("---")
        st.markdown("**계열별 비교**")
        all_ser_cmp = sorted(set(r["계열"] for rows in year_data.values() for r in rows if r["계열"]))
        ser_rows = []
        for ser in all_ser_cmp:
            row_d = {"계열": ser}
            for yr, rows in sorted(year_data.items(), reverse=True):
                fr = [r for r in rows if r["계열"] == ser
                      and (not sel_month_c or r["월"] in sel_month_c)]
                tot_정 = sum(r["정원"] for r in fr)
                tot_확 = sum(r["확정인원"] for r in fr)
                row_d[f"{yr}년 과정수"]   = len(fr)
                row_d[f"{yr}년 모집률(%)"] = round(tot_확/tot_정*100,1) if tot_정>0 else 0
            ser_rows.append(row_d)
        st.dataframe(pd.DataFrame(ser_rows), use_container_width=True, hide_index=True)

        # ── 과정별 상세 비교 ──────────────────────────
        st.markdown("---")
        st.markdown("**과정별 상세**")
        detail_yr = st.selectbox("연도 선택", sorted(year_data.keys(), reverse=True), key="det_yr")
        det_rows = filter_rows(year_data[detail_yr])
        if det_rows:
            st.dataframe(pd.DataFrame([{
                "월": r["월"], "계열": r["계열"], "지점": r["지점"],
                "과정명": r["과정명"], "시작일": r["시작일"], "종료일": r["종료일"],
                "정원": r["정원"], "확정": r["확정인원"], "신청": r["신청인원"],
                "모집률(%)": round(r["모집률"]*100,1),
                "신청률(%)": round(r["신청률"]*100,1),
            } for r in det_rows]), use_container_width=True, hide_index=True)
